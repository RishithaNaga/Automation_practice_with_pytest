import openpyxl

def getrowcount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.max_row

def getcolumncount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.max_column

def readdata(file,sheetName,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=colnum).value

def writeData(file,sheetName,rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=colnum).value=data
    workbook.save(file)



